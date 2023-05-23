from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests

# Função para obter os dados
def obter_dados_wikipedia(url):
    requisicao = requests.get(url)
    html = requisicao.text
    soup = BeautifulSoup(html, "html.parser")

    tabelas = soup.find_all("table", class_="wikitable")
    tabela_classificacao = None
    dados = []

    for tabela in tabelas:
        if "classificação" in tabela.get_text().lower():
            tabela_classificacao = tabela
            break

    if tabela_classificacao:
        linhas = tabela_classificacao.find_all("tr")[1:]
        for linha in linhas:
            colunas = linha.find_all("td")
            if len(colunas) >= 11:
                posicao = colunas[0].text.strip()
                time = colunas[1].text.strip()
                pontuacao = colunas[2].text.strip()
                aproveitamento = colunas[10].text.strip()

                dados.append([time, pontuacao, posicao, aproveitamento])
    
    return dados

# Dados da Wikipedia
url_wikipedia = "https://pt.wikipedia.org/wiki/Campeonato_Brasileiro_de_Futebol_de_2022_-_S%C3%A9rie_A"
dados_wikipedia = obter_dados_wikipedia(url_wikipedia)

# Excel
arquivo_excel = "E:/Campeonato_Brasileiro.xlsx"
planilha_desejada = "Campeonato 2022"
colunas = ["A", "B", "C", "D"]  # Colunas: "Times", "Pontuação", "Posição", "Aproveitamento"

# Função para inserir os dados no Excel
def inserir_dados_excel(arquivo, planilha, colunas, dados):
    workbook = load_workbook(arquivo)
    sheet = workbook[planilha]

    for i, coluna in enumerate(colunas):
        valores = [dados[j][i] for j in range(len(dados))]
        for j, valor in enumerate(valores):
            linha = j + 2
            if coluna == "D":  # Coluna do aproveitamento
                valor = f"{valor}%"
            sheet[f"{coluna}{linha}"] = valor

    workbook.save(arquivo)

def visualizar_colunas(arquivo, planilha, colunas):
    workbook = load_workbook(arquivo)
    sheet = workbook[planilha]

    for coluna in colunas:
        valores_coluna = [cell.value for cell in sheet[coluna][2:len(sheet[coluna]) + 1]]
        print(f"Coluna {coluna}:")
        for valor in valores_coluna:
            print(valor)
        print()


# Inserir os dados no Excel
try:
    inserir_dados_excel(arquivo_excel, planilha_desejada, colunas, dados_wikipedia)
    print("Inserção de dados bem-sucedida!")
except Exception as e:
    print("Ocorreu um erro durante a inserção de dados:", str(e))

# Visualizar as colunas após a inserção
print("Colunas após a inserção:")
visualizar_colunas(arquivo_excel, planilha_desejada, colunas)
